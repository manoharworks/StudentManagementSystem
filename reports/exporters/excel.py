"""
Generic Excel exporter.
"""

from io import BytesIO

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from reports.utils import get_field_value


class ExcelExporter:
    """
    Export any queryset as an Excel workbook.
    """

    @staticmethod
    def export(*, queryset, report):

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = report.title

        # -------------------------
        # Header
        # -------------------------

        for column_index, column in enumerate(
            report.columns,
            start=1,
        ):

            cell = worksheet.cell(
                row=1,
                column=column_index,
                value=column.header,
            )

            cell.font = Font(
                bold=True,
            )

        # -------------------------
        # Data
        # -------------------------

        row_number = 2

        for obj in queryset:

            column_number = 1

            for column in report.columns:

                if column.getter is not None:

                    value = column.getter(obj)

                else:

                    value = get_field_value(
                        obj,
                        column.field,
                    )

                worksheet.cell(
                    row=row_number,
                    column=column_number,
                    value=value,
                )

                column_number += 1

            row_number += 1

        # -------------------------
        # Auto Width
        # -------------------------

        for column_cells in worksheet.columns:

            max_length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            worksheet.column_dimensions[
                get_column_letter(
                    column_cells[0].column
                )
            ].width = max_length + 3

        # -------------------------
        # Download
        # -------------------------

        buffer = BytesIO()

        workbook.save(buffer)

        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/"
                "vnd.openxmlformats-"
                "officedocument."
                "spreadsheetml.sheet"
            ),
        )

        response["Content-Disposition"] = (
            f'attachment; filename="{report.filename}.xlsx"'
        )

        return response